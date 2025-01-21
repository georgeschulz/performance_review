import os
from pyairtable import Api
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import json
import datetime

###############################################################################
# STEP 1: Define a helper function to fetch and transform Airtable data into a DataFrame
###############################################################################
def fetch_airtable_data():
    """
    Fetch data from the specified Airtable base/table/view and return
    a Pandas DataFrame containing the columns required by channel_stats and close_rate.
    """
    # 1) Connect to Airtable
    api = Api(os.getenv('AIRTABLE_API_KEY'))

    # Update these to match your Base ID, Table ID, View name, etc.
    base_id = 'appBsty6iukfNnuEK'
    table_name = 'tblenVnxR8q8iTGSk'
    view_name = 'viwQMbNHcF8DMkZz8'

    table = api.table(base_id, table_name)
    records = table.all(view=view_name)

    # save the first record to a json file
    with open('first_record.json', 'w') as f:
        json.dump(records[0], f)

    # 2) Transform each record into a dict that aligns with the columns we need
    rows = []
    for record in records:
        f = record.get('fields', {})

        # Some fields might be missing or partial, so we use .get() with defaults
        # If “Salesperson” is a dict, we grab the 'name'
        salesperson = None
        if isinstance(f.get('Salesperson'), dict):
            salesperson = f['Salesperson'].get('name')

        # Extract close date in a format pandas can parse
        close_date = f.get('Close Date')
        # (Optionally convert the date string to consistent format if needed)

        # Decide which lead type field to use
        # If you prefer 'Lead Type (Clean)' over 'Lead Type', just adjust below
        lead_type = f.get('Lead Type (Clean)', f.get('Lead Type', 'Unknown'))

        row_data = {
            'Salesperson': salesperson,
            'Lead Type': lead_type,
            'Close Date': close_date,       # e.g. "2024-09-09"
            'Close Status': f.get('Close Status', None),
        }
        rows.append(row_data)
        

    # 3) Create a DataFrame
    df = pd.DataFrame(rows)

    # 4) Convert Close Date to datetime if it’s not already
    df['Close Date'] = pd.to_datetime(df['Close Date'], errors='coerce')

    return df

###############################################################################
# STEP 2: Adapted “channel_stats” function using Airtable data
###############################################################################
def channel_stats_staging():
    """
    Pull data from Airtable, generate channel stats, and save to an Excel file in 'outputs/'.
    """

    # 1) Get the DataFrame from Airtable
    df = fetch_airtable_data()

    # 2) Clean Lead Type
    df['Lead Type'] = df['Lead Type'].fillna('Unknown')
    
    # 3) Convert Close Date to datetime (fetch_airtable_data already does this) and add Month column
    df['Month'] = df['Close Date'].dt.strftime('%B %y')  # e.g. 'January 24'
    
    # 4) Initialize results list
    results = []
    
    # 5) Get unique months and lead types
    #    Sort the months in chronological order based on the datetime interpretation
    months = sorted(
        df['Month'].dropna().unique(),
        key=lambda x: pd.to_datetime(x, format='%B %y')
    )
    lead_types = df['Lead Type'].dropna().unique()

    # 6) Calculate metrics for each month and lead type
    for month in months:
        month_data = df[df['Month'] == month]
        row_data = {'Month': month}
        
        for lead_type in lead_types:
            type_data = month_data[month_data['Lead Type'] == lead_type]
            
            total_leads = len(type_data)
            disqualified = type_data['Close Status'].str.contains('Disqualified', na=False).sum()
            won_total = (
                (type_data['Close Status'] == 'Won: Recurring').sum() +
                (type_data['Close Status'] == 'Won: One Time').sum()
            )
            won_recurring = (type_data['Close Status'] == 'Won: Recurring').sum()
            
            lost = type_data['Close Status'].str.contains('Lost', na=False).sum()
            denominator = lost + won_total
            close_rate = (won_total / denominator * 100) if denominator > 0 else 0
            
            prefix = lead_type.replace(' ', '_')
            row_data.update({
                f'{prefix}_Total': total_leads,
                f'{prefix}_Disqualified': disqualified,
                f'{prefix}_Won': won_total,
                f'{prefix}_Won_Recurring': won_recurring,
                f'{prefix}_Close_Rate': f'{close_rate:.2f}%'
            })
        
        results.append(row_data)
    
    # 7) Convert results to DataFrame
    results_df = pd.DataFrame(results)
    
    # 8) Write final DataFrame to Excel with formatting
    with pd.ExcelWriter('staging/Channel Stats.xlsx', engine='openpyxl') as writer:
        # Create a multi-index column arrangement
        new_columns = pd.MultiIndex.from_tuples([('Month', '')])
        
        for lead_type in df['Lead Type'].unique():
            new_columns = new_columns.append(pd.MultiIndex.from_tuples([
                (lead_type, 'Total Leads'),
                (lead_type, 'Disqualified'),
                (lead_type, 'Won'),
                (lead_type, 'Won Recurring'),
                (lead_type, 'Close Rate')
            ]))
        
        formatted_df = pd.DataFrame(columns=new_columns)
        formatted_df['Month', ''] = results_df['Month']
        
        for lead_type in df['Lead Type'].unique():
            prefix = lead_type.replace(' ', '_')
            formatted_df[lead_type, 'Total Leads'] = results_df[f'{prefix}_Total']
            formatted_df[lead_type, 'Disqualified'] = results_df[f'{prefix}_Disqualified']
            formatted_df[lead_type, 'Won'] = results_df[f'{prefix}_Won']
            formatted_df[lead_type, 'Won Recurring'] = results_df[f'{prefix}_Won_Recurring']
            formatted_df[lead_type, 'Close Rate'] = results_df[f'{prefix}_Close_Rate']
        
        formatted_df.to_excel(writer, sheet_name='Channel Stats', index=True)
        
        workbook = writer.book
        worksheet = writer.sheets['Channel Stats']
        
        # Hide the index column (column A)
        worksheet.column_dimensions['A'].hidden = True
        
        # Define some colors
        colors = {
            'Thumbtack': 'E2EFDA',
            'Inbound': 'B7E1CD',
            'Unknown': 'D9D9D9',
            'Email Lead': 'C9DAF8',
            'Form Fill': 'D9EAD3',
            'Outbound': 'FFD966',
            'Sentricon Lead': 'FFF2CC',
            'Other Tech Lead': 'F4CCCC',
            'Referral': 'D0E0E3',
            'WTR Lead': 'FBE5D6',
            'WTR Free Trial Request': 'E6B8AF'
        }
        
        # Start formatting columns
        current_col = 2  # B (since A is hidden)
        
        # Skip the Month column group
        current_col += 1
        
        # Apply color to each lead type group
        for lead_type in df['Lead Type'].unique():
            start_col = current_col
            end_col = current_col + 4  # 5 columns for each lead type
            fill = PatternFill(
                start_color=colors.get(lead_type, 'D9D9D9'),
                end_color=colors.get(lead_type, 'D9D9D9'),
                fill_type='solid'
            )
            
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = fill
                cell.font = Font(bold=True, color='000000')
            
            current_col += 5
        
        # Adjust column widths
        for col in range(2, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 15
    
    print("Channel Stats Excel saved to staging/Channel Stats.xlsx")
    return results_df

###############################################################################
# STEP 3: Adapted “close_rate” function using Airtable data
###############################################################################
def close_rate_staging(salespeople=None, exclude_channels=None):
    """
    Pull data from Airtable, generate close rate stats, and save to a CSV in 'staging/'.
    """
    if salespeople is None:
        salespeople = []
    if exclude_channels is None:
        exclude_channels = []

    # 1) Get the DataFrame from Airtable
    df = fetch_airtable_data()

    # 2) Change: Use full date instead of month
    df['Day'] = df['Close Date'].dt.strftime('%Y-%m-%d')  # e.g. '2024-03-21'
    
    # 3) If no salespeople specified, use all unique
    if not salespeople:
        salespeople = df['Salesperson'].dropna().unique()
    
    results = []
    
    # 4) Calculate metrics for each salesperson and each day
    for sp in salespeople:
        sp_data = df[df['Salesperson'] == sp]

        # Exclude any leads from the given channels
        sp_data = sp_data[~sp_data['Lead Type'].isin(exclude_channels)]

        # Group by each day instead of month
        for day in sorted(sp_data['Day'].dropna().unique()):
            day_data = sp_data[sp_data['Day'] == day]
            
            lost = day_data['Close Status'].str.contains('Lost', na=False).sum()
            won_recurring = (day_data['Close Status'] == 'Won: Recurring').sum()
            won_one_time = (day_data['Close Status'] == 'Won: One Time').sum()
            disqualified = day_data['Close Status'].str.contains('Disqualified', na=False).sum()
            open_leads = day_data['Close Status'].isna().sum()
            
            denominator = lost + won_recurring + won_one_time
            close_rate = (won_recurring / denominator * 100) if denominator > 0 else 0
            
            total_closed = denominator + disqualified
            disqualify_rate = (disqualified / total_closed * 100) if total_closed > 0 else 0
            
            results.append({
                'Salesperson': sp,
                'Day': day,  # Changed from 'Month' to 'Day'
                'Lost': lost,
                'Won: Recurring': won_recurring,
                'Won: One Time': won_one_time,
                'Disqualified': disqualified,
                'Open': open_leads,
                'Close Rate': f'{close_rate:.2f}%',
                'Disqualify Rate': f'{disqualify_rate:.2f}%'
            })
    
    # 5) Convert results to DataFrame and save as CSV
    results_df = pd.DataFrame(results)
    results_df.to_csv('staging/Close Rate.csv', index=False)
    
    print("Close Rate CSV saved to staging/Close Rate.csv")
    return results_df

###############################################################################
# STEP 4: (Optional) Demo calls
###############################################################################
if __name__ == "__main__":
    # Run the adapted channel stats
    channel_df = channel_stats_staging()
    
    # Run the adapted close rate
    # Example usage, excluding some channel and specifying a couple of salespeople
    cr_df = close_rate_staging(
        salespeople=[],
        exclude_channels=["Unknown", "Test Channel"]
    )
