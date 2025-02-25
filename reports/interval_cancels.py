import pandas as pd
import datetime
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font
import os

def get_fiscal_year_start(current_date=None):
    """
    Calculate the start date of the current fiscal year (September 1st)
    """
    if current_date is None:
        current_date = datetime.datetime.now()
    
    # If current date is before September 1st, fiscal year started last year
    if current_date.month < 9:
        return datetime.datetime(current_date.year - 1, 9, 1)
    # Otherwise fiscal year started this year
    else:
        return datetime.datetime(current_date.year, 9, 1)

def get_month_start(current_date=None):
    """
    Calculate the start date of the current month
    """
    if current_date is None:
        current_date = datetime.datetime.now()
    
    return datetime.datetime(current_date.year, current_date.month, 1)

def calculate_data(data, date_column):
    """
    Calculate data from a dataframe based on the specified date column
    """
    count = len(data)
    
    return {
        'Count': count
    }

def interval_cancels(beginning_of_time=None, salespeople=None, current_date=None):
    """
    Generate interval-based cancellation and start reports by salesperson:
    1. Year-to-date and month-to-date cancellations and starts
    2. Weekly cancellations and starts (Monday to Sunday) with MTD and YTD metrics for each week
    
    Parameters:
    - beginning_of_time: datetime or string, minimum date to include data from
    - salespeople: list of salespeople to include (defaults to all salespeople)
    - current_date: datetime, date to use as "now" (defaults to today)
    """
    # Set current date if not provided
    if current_date is None:
        current_date = datetime.datetime.now()
    elif isinstance(current_date, str):
        current_date = pd.to_datetime(current_date)
    
    # Ensure weekly_outputs directory exists
    os.makedirs('weekly_outputs', exist_ok=True)
    
    # Read the CSV file
    df = pd.read_csv('weekly_review_data/Starts.csv')
    
    # Convert date columns to datetime
    df['Cancel Date'] = pd.to_datetime(df['Cancel Date'], errors='coerce')
    df['Date Added'] = pd.to_datetime(df['Date Added'], errors='coerce')
    
    # Handle empty Salesperson column by setting it to "Other Rep"
    df['Effective Salesperson'] = df['Salesperson'].fillna("Other Rep")
    
    # If salespeople list is provided, filter to only those salespeople
    if salespeople:
        # Create a category for sales not belonging to the specified salespeople
        df.loc[~df['Effective Salesperson'].isin(salespeople), 'Effective Salesperson'] = "Other Rep"
        unique_salespeople = list(salespeople) + ["Other Rep"]
    else:
        # Get all unique salespeople
        unique_salespeople = df['Effective Salesperson'].unique().tolist()
    
    # Filter by beginning_of_time if provided
    if beginning_of_time:
        if isinstance(beginning_of_time, str):
            beginning_of_time = pd.to_datetime(beginning_of_time)
        # Filter both cancel and date added dates
        df = df[((df['Cancel Date'] >= beginning_of_time) & (~df['Cancel Date'].isna())) | 
                (df['Date Added'] >= beginning_of_time)]
    
    # Calculate fiscal year and month start dates
    fiscal_year_start = get_fiscal_year_start(current_date)
    month_start = get_month_start(current_date)
    
    # Initialize results for YTD and MTD report
    ytd_mtd_results = []
    
    # Calculate YTD and MTD metrics for each salesperson
    for sp in unique_salespeople:
        sp_data = df[df['Effective Salesperson'] == sp]
        
        # Year-to-date cancellations
        ytd_cancels = sp_data[(sp_data['Cancel Date'] >= fiscal_year_start) & 
                             (sp_data['Cancel Date'] <= current_date)]
        ytd_cancel_metrics = calculate_data(ytd_cancels, 'Cancel Date')
        
        # Month-to-date cancellations
        mtd_cancels = sp_data[(sp_data['Cancel Date'] >= month_start) & 
                             (sp_data['Cancel Date'] <= current_date)]
        mtd_cancel_metrics = calculate_data(mtd_cancels, 'Cancel Date')
        
        # Year-to-date starts
        ytd_starts = sp_data[(sp_data['Date Added'] >= fiscal_year_start) & 
                            (sp_data['Date Added'] <= current_date)]
        ytd_start_metrics = calculate_data(ytd_starts, 'Date Added')
        
        # Month-to-date starts
        mtd_starts = sp_data[(sp_data['Date Added'] >= month_start) & 
                            (sp_data['Date Added'] <= current_date)]
        mtd_start_metrics = calculate_data(mtd_starts, 'Date Added')
        
        # Add to results
        ytd_mtd_results.append({
            'Salesperson': sp,
            'YTD Cancels': ytd_cancel_metrics['Count'],
            'MTD Cancels': mtd_cancel_metrics['Count'],
            'YTD Starts': ytd_start_metrics['Count'],
            'MTD Starts': mtd_start_metrics['Count']
        })
    
    # Convert YTD/MTD results to DataFrame
    ytd_mtd_df = pd.DataFrame(ytd_mtd_results)
    
    # Initialize results for mega report (weekly + MTD + YTD)
    mega_report_results = []
    
    # Find the Monday of the week containing the fiscal year start date
    # This ensures we start on a Monday
    days_since_monday = fiscal_year_start.weekday()
    start_monday = fiscal_year_start - timedelta(days=days_since_monday)
    
    # Generate weekly periods from start_monday to current_date
    current_monday = start_monday
    while current_monday <= current_date:
        week_end = current_monday + timedelta(days=6)  # Sunday
        
        # If week_end is beyond current_date, cap it at current_date
        week_end = min(week_end, current_date)
        
        # Get the month start for this specific week
        week_month_start = get_month_start(week_end)
        
        # Calculate weekly metrics for each salesperson
        for sp in unique_salespeople:
            sp_data = df[df['Effective Salesperson'] == sp]
            
            # Weekly cancellations (just for this week)
            weekly_cancels = sp_data[(sp_data['Cancel Date'] >= current_monday) & 
                                   (sp_data['Cancel Date'] <= week_end)]
            
            # Weekly starts (just for this week)
            weekly_starts = sp_data[(sp_data['Date Added'] >= current_monday) & 
                                  (sp_data['Date Added'] <= week_end)]
            
            # Only add to results if there's any data for the week
            if len(weekly_cancels) > 0 or len(weekly_starts) > 0:
                weekly_cancel_metrics = calculate_data(weekly_cancels, 'Cancel Date')
                weekly_start_metrics = calculate_data(weekly_starts, 'Date Added')
                
                # YTD cancellations up to and including the end of this week
                ytd_cancels = sp_data[(sp_data['Cancel Date'] >= fiscal_year_start) & 
                                    (sp_data['Cancel Date'] <= week_end)]
                ytd_cancel_metrics = calculate_data(ytd_cancels, 'Cancel Date')
                
                # MTD cancellations up to and including the end of this week
                mtd_cancels = sp_data[(sp_data['Cancel Date'] >= week_month_start) & 
                                    (sp_data['Cancel Date'] <= week_end)]
                mtd_cancel_metrics = calculate_data(mtd_cancels, 'Cancel Date')
                
                # YTD starts up to and including the end of this week
                ytd_starts = sp_data[(sp_data['Date Added'] >= fiscal_year_start) & 
                                   (sp_data['Date Added'] <= week_end)]
                ytd_start_metrics = calculate_data(ytd_starts, 'Date Added')
                
                # MTD starts up to and including the end of this week
                mtd_starts = sp_data[(sp_data['Date Added'] >= week_month_start) & 
                                   (sp_data['Date Added'] <= week_end)]
                mtd_start_metrics = calculate_data(mtd_starts, 'Date Added')
                
                mega_report_results.append({
                    'Salesperson': sp,
                    'Week Start': current_monday.strftime('%Y-%m-%d'),
                    'Week End': week_end.strftime('%Y-%m-%d'),
                    # Weekly metrics
                    'Weekly Cancels': weekly_cancel_metrics['Count'],
                    'Weekly Starts': weekly_start_metrics['Count'],
                    # MTD metrics as of this week (including this week)
                    'MTD Cancels': mtd_cancel_metrics['Count'],
                    'MTD Starts': mtd_start_metrics['Count'],
                    # YTD metrics as of this week (including this week)
                    'YTD Cancels': ytd_cancel_metrics['Count'],
                    'YTD Starts': ytd_start_metrics['Count']
                })
        
        # Move to next Monday
        current_monday += timedelta(days=7)
    
    # Convert mega report results to DataFrame
    mega_report_df = pd.DataFrame(mega_report_results)
    
    # Save as Excel with formatting
    excel_path = 'weekly_outputs/Cancel Report.xlsx'
    
    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Write each dataframe to a different worksheet
        mega_report_df.to_excel(writer, sheet_name='Weekly Report', index=False)
        ytd_mtd_df.to_excel(writer, sheet_name='YTD and MTD Summary', index=False)
    
    # Apply formatting to the Excel file
    workbook = openpyxl.load_workbook(excel_path)
    
    # Format the Weekly Report sheet
    worksheet = workbook['Weekly Report']
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        # Find the maximum length in the column
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Set the column width (with some padding)
        adjusted_width = max_length + 4
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Find the column indices for cancel columns only
    cancel_columns = []
    for col_idx, col_name in enumerate(mega_report_df.columns, start=1):
        if 'Cancels' in col_name:
            cancel_columns.append(col_idx)
    
    # Apply bold formatting to the cancel columns only
    bold_font = Font(bold=True)
    for col_idx in cancel_columns:
        # Bold the header
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.font = bold_font
        
        # Bold all data cells in the column
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = bold_font
    
    # Format the YTD and MTD Summary sheet
    worksheet = workbook['YTD and MTD Summary']
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        # Find the maximum length in the column
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Set the column width (with some padding)
        adjusted_width = max_length + 4
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Find the column indices for cancel columns only
    cancel_columns = []
    for col_idx, col_name in enumerate(ytd_mtd_df.columns, start=1):
        if 'Cancels' in col_name:
            cancel_columns.append(col_idx)
    
    # Apply bold formatting to the cancel columns only
    for col_idx in cancel_columns:
        # Bold the header
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.font = bold_font
        
        # Bold all data cells in the column
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = bold_font
    
    # Save the formatted workbook
    workbook.save(excel_path)
    
    # Create weekly dataframe for return value
    weekly_results = []
    current_monday = start_monday
    while current_monday <= current_date:
        week_end = current_monday + timedelta(days=6)  # Sunday
        week_end = min(week_end, current_date)
        
        for sp in unique_salespeople:
            sp_data = df[df['Effective Salesperson'] == sp]
            
            weekly_cancels = sp_data[(sp_data['Cancel Date'] >= current_monday) & 
                                   (sp_data['Cancel Date'] <= week_end)]
            
            weekly_starts = sp_data[(sp_data['Date Added'] >= current_monday) & 
                                  (sp_data['Date Added'] <= week_end)]
            
            if len(weekly_cancels) > 0 or len(weekly_starts) > 0:
                weekly_cancel_metrics = calculate_data(weekly_cancels, 'Cancel Date')
                weekly_start_metrics = calculate_data(weekly_starts, 'Date Added')
                
                weekly_results.append({
                    'Salesperson': sp,
                    'Week Start': current_monday.strftime('%Y-%m-%d'),
                    'Week End': week_end.strftime('%Y-%m-%d'),
                    'Cancels': weekly_cancel_metrics['Count'],
                    'Starts': weekly_start_metrics['Count']
                })
        
        current_monday += timedelta(days=7)
    
    weekly_df = pd.DataFrame(weekly_results)
    
    return {
        'ytd_mtd': ytd_mtd_df,
        'weekly': weekly_df,
        'mega_report': mega_report_df
    }

if __name__ == "__main__":
    # Run the report with default parameters
    interval_cancels()
