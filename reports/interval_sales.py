import pandas as pd
import datetime
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font
import os

def get_fiscal_year_start(current_date=None):
    """
    Calculate the start date of the current fiscal year (September 1st)
    """
    if current_date is None:
        current_date = datetime.datetime.now()
    
    # If current date is before September 1st, fiscal year started last year
    if current_date.month < 9:
        return datetime.datetime(current_date.year - 1, 9, 1)
    # Otherwise fiscal year started this year
    else:
        return datetime.datetime(current_date.year, 9, 1)

def get_month_start(current_date=None):
    """
    Calculate the start date of the current month
    """
    if current_date is None:
        current_date = datetime.datetime.now()
    
    return datetime.datetime(current_date.year, current_date.month, 1)

def calculate_sales_data(data):
    """
    Calculate sales data from a dataframe of sales, subtracting Amount Before from First Year ACV
    """
    # Convert First Year ACV and Amount Before to numeric, removing $ and commas
    sales_values = data['First Year ACV'].replace('[\$,]', '', regex=True).astype(float)
    amount_before = data['Amount Before'].astype(float)
    
    # Calculate net sales (First Year ACV - Amount Before)
    net_sales = sales_values - amount_before
    
    total_sales = net_sales.sum()
    count = len(data)
    
    return {
        'Total Sales': total_sales,
        'Count': count,
        'Average Sale': total_sales / count if count > 0 else 0
    }

def generate_mtd_drill_down(df, salespeople, current_date=None):
    """
    Generate MTD drill-down reports for each rep with specific columns
    
    Parameters:
    - df: DataFrame containing the sales data
    - salespeople: list of salespeople to generate reports for
    - current_date: datetime, date to use as "now" (defaults to today)
    """
    if current_date is None:
        current_date = datetime.datetime.now()
    
    # Ensure weekly_outputs directory exists
    os.makedirs('weekly_outputs', exist_ok=True)
    
    # Get the start of the current month
    month_start = get_month_start(current_date)
    
    # For each salesperson, create a drill-down report
    for rep in salespeople:
        # Filter data for this rep and MTD
        rep_data = df[(df['Salesperson'] == rep) & 
                      (df['Add Date'] >= month_start) & 
                      (df['Add Date'] <= current_date)]
        
        # Select columns, now including Amount Before
        drill_down_data = rep_data[['Location', 'First Name', 'Last Name', 
                                   'Service Code', 'Add Date', 'Start Date', 
                                   'First Year ACV', 'Amount Before']]
        
        # Sort by Add Date
        drill_down_data = drill_down_data.sort_values(by='Add Date')
        
        # Calculate totals
        acv_total = drill_down_data['First Year ACV'].replace('[\$,]', '', regex=True).astype(float).sum()
        amount_before_total = drill_down_data['Amount Before'].astype(float).sum()
        net_total = acv_total - amount_before_total
        
        # Create a filename with the rep's name
        rep_name_for_file = rep.replace(' ', '_').lower()
        file_path = f'weekly_outputs/{rep_name_for_file}_mtd_drill_down.xlsx'
        
        # Save to Excel
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            drill_down_data.to_excel(writer, index=False, sheet_name='MTD Drill Down')
            
            workbook = writer.book
            worksheet = writer.sheets['MTD Drill Down']
            
            # Add total rows
            row_num = len(drill_down_data) + 2  # +2 for header and 1-based indexing
            
            # Add First Year ACV total
            worksheet.cell(row=row_num, column=1, value="TOTAL First Year ACV")
            worksheet.cell(row=row_num, column=7, value=acv_total)
            worksheet.cell(row=row_num, column=7).number_format = '$#,##0.00'
            
            # Add Amount Before total
            worksheet.cell(row=row_num + 1, column=1, value="TOTAL Amount Before")
            worksheet.cell(row=row_num + 1, column=8, value=amount_before_total)
            worksheet.cell(row=row_num + 1, column=8).number_format = '$#,##0.00'
            
            # Add Net total
            worksheet.cell(row=row_num + 2, column=1, value="NET TOTAL")
            worksheet.cell(row=row_num + 2, column=7, value=net_total)
            worksheet.cell(row=row_num + 2, column=7).number_format = '$#,##0.00'
            
            # Format the currency columns
            for row_idx in range(2, row_num):  # Start from 2 to skip header
                # Format First Year ACV column
                cell = worksheet.cell(row=row_idx, column=7)
                cell.number_format = '$#,##0.00'
                
                # Format Amount Before column
                cell = worksheet.cell(row=row_idx, column=8)
                cell.number_format = '$#,##0.00'
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = max_length + 4
                worksheet.column_dimensions[column_letter].width = adjusted_width

def calculate_amount_before(price_analysis_df):
    """
    Calculate the Amount Before column based on cancelled setups from Starts.csv
    """
    # Read the Starts.csv file
    starts_df = pd.read_csv('weekly_review_data/Starts.csv')
    
    # Convert date columns to datetime
    starts_df['Cancel Date'] = pd.to_datetime(starts_df['Cancel Date'], errors='coerce')
    price_analysis_df['Start Date'] = pd.to_datetime(price_analysis_df['Start Date'], errors='coerce')
    
    # Initialize Amount Before column
    price_analysis_df['Amount Before'] = 0.0
    
    # Iterate through Price Analysis rows
    for idx, row in price_analysis_df.iterrows():
        location = row['Location']  # Changed from Location Code to Location
        setup_id = row['SetupID']
        start_date = row['Start Date']
        
        if pd.isna(start_date):
            continue
            
        # Find relevant cancellations in Starts df
        mask = (
            (starts_df['Location Code'] == location) &  # Starts.csv uses Location Code
            (starts_df['SetupID'] != setup_id) &
            (~starts_df['Cancel Date'].isna()) &
            (abs((starts_df['Cancel Date'] - start_date).dt.days) <= 30)
        )
        
        # Sum the Recurring Value for matching rows
        amount_before = starts_df.loc[mask, 'Recurring Value'].sum()
        price_analysis_df.at[idx, 'Amount Before'] = amount_before
    
    return price_analysis_df

def interval_sales(beginning_of_time=None, salespeople=["Hussam Olabi", "Kamaal Sherrod", "Rob Dively"], exclude_sale_types=[], current_date=None):
    """
    Generate interval-based sales reports:
    1. Year-to-date and month-to-date sales
    2. Weekly sales (Monday to Sunday) with MTD and YTD metrics for each week
    3. MTD drill-down reports for each rep with specific columns
    
    Parameters:
    - beginning_of_time: datetime or string, minimum date to include sales from
    - salespeople: list of salespeople to include (defaults to the three specified reps)
    - exclude_sale_types: list of sale types to exclude
    - current_date: datetime, date to use as "now" (defaults to today)
    """
    # Set current date if not provided
    if current_date is None:
        current_date = datetime.datetime.now()
    elif isinstance(current_date, str):
        current_date = pd.to_datetime(current_date)
    
    # Ensure weekly_outputs directory exists
    os.makedirs('weekly_outputs', exist_ok=True)
    
    # Read the CSV file
    df = pd.read_csv('weekly_review_data/Price Analysis.csv')
    
    # Calculate Amount Before and save transformed file
    df = calculate_amount_before(df)
    df.to_csv('weekly_review_data/Price Analysis_transformed.csv', index=False)
    
    # Convert Add Date to datetime
    df['Add Date'] = pd.to_datetime(df['Add Date'], errors='coerce')
    
    # Filter by beginning_of_time if provided
    if beginning_of_time:
        if isinstance(beginning_of_time, str):
            beginning_of_time = pd.to_datetime(beginning_of_time)
        df = df[df['Add Date'] >= beginning_of_time]
    
    # Handle empty Salesperson column by setting it to "Other Rep"
    df['Effective Salesperson'] = df['Salesperson'].fillna("Other Rep")
    
    # Create a category for sales not belonging to the specified salespeople
    df.loc[~df['Effective Salesperson'].isin(salespeople), 'Effective Salesperson'] = "Other Rep"
    
    # Get unique salespeople (the specified ones plus "Other Rep")
    unique_salespeople = list(salespeople) + ["Other Rep"]
    
    # Exclude specified sale types
    if exclude_sale_types:
        df = df[~df['Sale Type'].isin(exclude_sale_types)]
    
    # Calculate fiscal year and month start dates
    fiscal_year_start = get_fiscal_year_start(current_date)
    
    # Initialize results for YTD and MTD report
    ytd_mtd_results = []
    
    # Calculate YTD and MTD metrics for each salesperson
    for sp in unique_salespeople:
        sp_data = df[df['Effective Salesperson'] == sp]
        
        # Year-to-date data
        ytd_data = sp_data[(sp_data['Add Date'] >= fiscal_year_start) & 
                           (sp_data['Add Date'] <= current_date)]
        ytd_metrics = calculate_sales_data(ytd_data)
        
        # Month-to-date data
        month_start = get_month_start(current_date)
        mtd_data = sp_data[(sp_data['Add Date'] >= month_start) & 
                           (sp_data['Add Date'] <= current_date)]
        mtd_metrics = calculate_sales_data(mtd_data)
        
        # Add to results
        ytd_mtd_results.append({
            'Salesperson': sp,
            'YTD Total Sales': ytd_metrics['Total Sales'],
            'YTD Count': ytd_metrics['Count'],
            'YTD Average Sale': ytd_metrics['Average Sale'],
            'MTD Total Sales': mtd_metrics['Total Sales'],
            'MTD Count': mtd_metrics['Count'],
            'MTD Average Sale': mtd_metrics['Average Sale']
        })
    
    # Convert YTD/MTD results to DataFrame
    ytd_mtd_df = pd.DataFrame(ytd_mtd_results)
    
    # Initialize results for mega report (weekly + MTD + YTD)
    mega_report_results = []
    
    # Find the Monday of the week containing the fiscal year start date
    # This ensures we start on a Monday
    days_since_monday = fiscal_year_start.weekday()
    start_monday = fiscal_year_start - timedelta(days=days_since_monday)
    
    # Generate weekly periods from start_monday to current_date
    current_monday = start_monday
    while current_monday <= current_date:
        week_end = current_monday + timedelta(days=6)  # Sunday
        
        # If week_end is beyond current_date, cap it at current_date
        week_end = min(week_end, current_date)
        
        # Get the month start for this specific week
        week_month_start = get_month_start(week_end)
        
        # Calculate weekly metrics for each salesperson
        for sp in unique_salespeople:
            sp_data = df[df['Effective Salesperson'] == sp]
            
            # Weekly data (just for this week)
            weekly_data = sp_data[(sp_data['Add Date'] >= current_monday) & 
                                 (sp_data['Add Date'] <= week_end)]
            
            # Only add to results if there's any data for the week
            if len(weekly_data) > 0:
                weekly_metrics = calculate_sales_data(weekly_data)
                
                # YTD data up to and including the end of this week
                ytd_data = sp_data[(sp_data['Add Date'] >= fiscal_year_start) & 
                                  (sp_data['Add Date'] <= week_end)]
                ytd_metrics = calculate_sales_data(ytd_data)
                
                # MTD data up to and including the end of this week
                mtd_data = sp_data[(sp_data['Add Date'] >= week_month_start) & 
                                  (sp_data['Add Date'] <= week_end)]
                mtd_metrics = calculate_sales_data(mtd_data)
                
                mega_report_results.append({
                    'Salesperson': sp,
                    'Week Start': current_monday.strftime('%Y-%m-%d'),
                    'Week End': week_end.strftime('%Y-%m-%d'),
                    # Weekly metrics
                    'Weekly Total Sales': weekly_metrics['Total Sales'],
                    'Weekly Count': weekly_metrics['Count'],
                    'Weekly Average Sale': weekly_metrics['Average Sale'],
                    # MTD metrics as of this week (including this week)
                    'MTD Total Sales': mtd_metrics['Total Sales'],
                    'MTD Count': mtd_metrics['Count'],
                    'MTD Average Sale': mtd_metrics['Average Sale'],
                    # YTD metrics as of this week (including this week)
                    'YTD Total Sales': ytd_metrics['Total Sales'],
                    'YTD Count': ytd_metrics['Count'],
                    'YTD Average Sale': ytd_metrics['Average Sale']
                })
        
        # Move to next Monday
        current_monday += timedelta(days=7)
    
    # Convert mega report results to DataFrame
    mega_report_df = pd.DataFrame(mega_report_results)
    
    # Save as Excel with formatting
    excel_path = 'weekly_outputs/Sales Report.xlsx'
    mega_report_df.to_excel(excel_path, index=False, engine='openpyxl')
    
    # Apply formatting to the Excel file
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook.active
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        # Find the maximum length in the column
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Set the column width (with some padding)
        adjusted_width = max_length + 4
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Find the column indices for sales and average columns
    currency_columns = []
    for col_idx, col_name in enumerate(mega_report_df.columns, start=1):
        if 'Total Sales' in col_name or 'Average Sale' in col_name:
            currency_columns.append(col_idx)
    
    # Apply bold formatting and currency format to the sales columns
    bold_font = Font(bold=True)
    for col_idx in currency_columns:
        # Bold the header
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.font = bold_font
        
        # Bold all data cells in the column and apply currency format
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = bold_font
            cell.number_format = '$#,##0.00'  # Format as currency with 2 decimal places
    
    # Save the formatted workbook
    workbook.save(excel_path)
    
    # Create weekly dataframe for return value (but don't save to CSV)
    weekly_results = []
    current_monday = start_monday
    while current_monday <= current_date:
        week_end = current_monday + timedelta(days=6)  # Sunday
        week_end = min(week_end, current_date)
        
        for sp in unique_salespeople:
            sp_data = df[df['Effective Salesperson'] == sp]
            weekly_data = sp_data[(sp_data['Add Date'] >= current_monday) & 
                                 (sp_data['Add Date'] <= week_end)]
            
            if len(weekly_data) > 0:
                weekly_metrics = calculate_sales_data(weekly_data)
                
                weekly_results.append({
                    'Salesperson': sp,
                    'Week Start': current_monday.strftime('%Y-%m-%d'),
                    'Week End': week_end.strftime('%Y-%m-%d'),
                    'Total Sales': weekly_metrics['Total Sales'],
                    'Count': weekly_metrics['Count'],
                    'Average Sale': weekly_metrics['Average Sale']
                })
        
        current_monday += timedelta(days=7)
    
    weekly_df = pd.DataFrame(weekly_results)
    
    # Generate MTD drill-down reports for each salesperson
    # Only generate for the actual salespeople (not "Other Rep")
    generate_mtd_drill_down(df, salespeople, current_date)
    
    return {
        'ytd_mtd': ytd_mtd_df,
        'weekly': weekly_df,
        'mega_report': mega_report_df
    }

if __name__ == "__main__":
    # Run the report with default parameters
    interval_sales() 