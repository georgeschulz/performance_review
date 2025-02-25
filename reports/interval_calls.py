import pandas as pd
import datetime
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font
import os

def get_fiscal_year_start(current_date=None):
    """
    Calculate the start date of the current fiscal year (September 1st)
    """
    if current_date is None:
        current_date = datetime.datetime.now()
    
    # If current date is before September 1st, fiscal year started last year
    if current_date.month < 9:
        return datetime.datetime(current_date.year - 1, 9, 1)
    # Otherwise fiscal year started this year
    else:
        return datetime.datetime(current_date.year, 9, 1)

def get_month_start(current_date=None):
    """
    Calculate the start date of the current month
    """
    if current_date is None:
        current_date = datetime.datetime.now()
    
    return datetime.datetime(current_date.year, current_date.month, 1)

def calculate_call_data(data):
    """
    Calculate call metrics from a dataframe of calls
    """
    # Count total calls
    total_calls = len(data)
    
    # Calculate total talk time (in seconds)
    # First, convert the talk time strings to seconds
    talk_times = []
    for talk_time in data['Talk Time']:
        if pd.isna(talk_time) or talk_time == '':
            talk_times.append(0)
        else:
            try:
                # Format is HH:MM:SS
                parts = talk_time.split(':')
                seconds = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
                talk_times.append(seconds)
            except (ValueError, IndexError):
                talk_times.append(0)
    
    total_talk_time_seconds = sum(talk_times)
    
    # Calculate average talk time
    avg_talk_time_seconds = total_talk_time_seconds / total_calls if total_calls > 0 else 0
    
    # Count answered calls
    answered_calls = len(data[data['Call Status'] == 'answered'])
    
    # Calculate answer rate
    answer_rate = answered_calls / total_calls if total_calls > 0 else 0
    
    return {
        'Total Calls': total_calls,
        'Answered Calls': answered_calls,
        'Answer Rate': answer_rate,
        'Total Talk Time (seconds)': total_talk_time_seconds,
        'Average Talk Time (seconds)': avg_talk_time_seconds
    }

def format_time_seconds(seconds):
    """
    Format seconds as HH:MM:SS
    """
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    seconds = int(seconds % 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

def interval_calls(beginning_of_time=None, agents=None, exclude_call_statuses=[], replacements={}, current_date=None):
    """
    Generate interval-based call reports:
    1. Year-to-date and month-to-date calls
    2. Weekly calls (Monday to Sunday) with MTD and YTD metrics for each week
    
    Parameters:
    - beginning_of_time: datetime or string, minimum date to include calls from
    - agents: list of agents to include (defaults to all agents)
    - exclude_call_statuses: list of call statuses to exclude
    - current_date: datetime, date to use as "now" (defaults to today)
    """
    # Set current date if not provided
    if current_date is None:
        current_date = datetime.datetime.now()
    elif isinstance(current_date, str):
        current_date = pd.to_datetime(current_date)
    
    # Ensure weekly_outputs directory exists
    os.makedirs('weekly_outputs', exist_ok=True)
    
    # Read the CSV file
    df = pd.read_csv('weekly_review_data/Calls.csv')
    
    # Convert Date to datetime
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    
    # Filter by beginning_of_time if provided
    if beginning_of_time:
        if isinstance(beginning_of_time, str):
            beginning_of_time = pd.to_datetime(beginning_of_time)
        df = df[df['Date'] >= beginning_of_time]
    
    # Handle empty Agent column by setting it to "Unassigned"
    df['Effective Agent'] = df['Agent'].fillna("Unassigned")
    
    # Apply replacements to convert email addresses to names
    if replacements:
        df['Effective Agent'] = df['Effective Agent'].map(lambda x: replacements.get(x, x))
    
    # Get unique agents
    if agents is None:
        # Use all agents in the data
        unique_agents = df['Effective Agent'].unique().tolist()
    else:
        # Use specified agents and add "Unassigned" for calls without an agent
        unique_agents = list(agents)
        if "Unassigned" not in unique_agents:
            unique_agents.append("Unassigned")
    
    # Exclude specified call statuses
    if exclude_call_statuses:
        df = df[~df['Call Status'].isin(exclude_call_statuses)]
    
    # Calculate fiscal year and month start dates
    fiscal_year_start = get_fiscal_year_start(current_date)
    
    # Initialize results for YTD and MTD report
    ytd_mtd_results = []
    
    # Calculate YTD and MTD metrics for each agent
    for agent in unique_agents:
        agent_data = df[df['Effective Agent'] == agent]
        
        # Year-to-date data
        ytd_data = agent_data[(agent_data['Date'] >= fiscal_year_start) & 
                             (agent_data['Date'] <= current_date)]
        ytd_metrics = calculate_call_data(ytd_data)
        
        # Month-to-date data
        month_start = get_month_start(current_date)
        mtd_data = agent_data[(agent_data['Date'] >= month_start) & 
                             (agent_data['Date'] <= current_date)]
        mtd_metrics = calculate_call_data(mtd_data)
        
        # Add to results
        ytd_mtd_results.append({
            'Agent': agent,
            'YTD Total Calls': ytd_metrics['Total Calls'],
            'YTD Answered Calls': ytd_metrics['Answered Calls'],
            'YTD Answer Rate': ytd_metrics['Answer Rate'],
            'YTD Total Talk Time': ytd_metrics['Total Talk Time (seconds)'],
            'YTD Average Talk Time': ytd_metrics['Average Talk Time (seconds)'],
            'MTD Total Calls': mtd_metrics['Total Calls'],
            'MTD Answered Calls': mtd_metrics['Answered Calls'],
            'MTD Answer Rate': mtd_metrics['Answer Rate'],
            'MTD Total Talk Time': mtd_metrics['Total Talk Time (seconds)'],
            'MTD Average Talk Time': mtd_metrics['Average Talk Time (seconds)']
        })
    
    # Convert YTD/MTD results to DataFrame
    ytd_mtd_df = pd.DataFrame(ytd_mtd_results)
    
    # Initialize results for mega report (weekly + MTD + YTD)
    mega_report_results = []
    
    # Find the Monday of the week containing the fiscal year start date
    # This ensures we start on a Monday
    days_since_monday = fiscal_year_start.weekday()
    start_monday = fiscal_year_start - timedelta(days=days_since_monday)
    
    # Generate weekly periods from start_monday to current_date
    current_monday = start_monday
    while current_monday <= current_date:
        week_end = current_monday + timedelta(days=6)  # Sunday
        
        # If week_end is beyond current_date, cap it at current_date
        week_end = min(week_end, current_date)
        
        # Get the month start for this specific week
        week_month_start = get_month_start(week_end)
        
        # Calculate weekly metrics for each agent
        for agent in unique_agents:
            agent_data = df[df['Effective Agent'] == agent]
            
            # Weekly data (just for this week)
            weekly_data = agent_data[(agent_data['Date'] >= current_monday) & 
                                   (agent_data['Date'] <= week_end)]
            
            # Only add to results if there's any data for the week
            if len(weekly_data) > 0:
                weekly_metrics = calculate_call_data(weekly_data)
                
                # YTD data up to and including the end of this week
                ytd_data = agent_data[(agent_data['Date'] >= fiscal_year_start) & 
                                    (agent_data['Date'] <= week_end)]
                ytd_metrics = calculate_call_data(ytd_data)
                
                # MTD data up to and including the end of this week
                mtd_data = agent_data[(agent_data['Date'] >= week_month_start) & 
                                    (agent_data['Date'] <= week_end)]
                mtd_metrics = calculate_call_data(mtd_data)
                
                mega_report_results.append({
                    'Agent': agent,
                    'Week Start': current_monday.strftime('%Y-%m-%d'),
                    'Week End': week_end.strftime('%Y-%m-%d'),
                    # Weekly metrics
                    'Weekly Total Calls': weekly_metrics['Total Calls'],
                    'Weekly Answered Calls': weekly_metrics['Answered Calls'],
                    'Weekly Answer Rate': weekly_metrics['Answer Rate'],
                    'Weekly Total Talk Time': weekly_metrics['Total Talk Time (seconds)'],
                    'Weekly Average Talk Time': weekly_metrics['Average Talk Time (seconds)'],
                    # MTD metrics as of this week (including this week)
                    'MTD Total Calls': mtd_metrics['Total Calls'],
                    'MTD Answered Calls': mtd_metrics['Answered Calls'],
                    'MTD Answer Rate': mtd_metrics['Answer Rate'],
                    'MTD Total Talk Time': mtd_metrics['Total Talk Time (seconds)'],
                    'MTD Average Talk Time': mtd_metrics['Average Talk Time (seconds)'],
                    # YTD metrics as of this week (including this week)
                    'YTD Total Calls': ytd_metrics['Total Calls'],
                    'YTD Answered Calls': ytd_metrics['Answered Calls'],
                    'YTD Answer Rate': ytd_metrics['Answer Rate'],
                    'YTD Total Talk Time': ytd_metrics['Total Talk Time (seconds)'],
                    'YTD Average Talk Time': ytd_metrics['Average Talk Time (seconds)']
                })
        
        # Move to next Monday
        current_monday += timedelta(days=7)
    
    # Convert mega report results to DataFrame
    mega_report_df = pd.DataFrame(mega_report_results)
    
    # Save as Excel with formatting
    excel_path = 'weekly_outputs/Calls Report.xlsx'
    mega_report_df.to_excel(excel_path, index=False, engine='openpyxl')
    
    # Apply formatting to the Excel file
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook.active
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        # Find the maximum length in the column
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Set the column width (with some padding)
        adjusted_width = max_length + 4
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Find the column indices for time and rate columns
    time_columns = []
    rate_columns = []
    for col_idx, col_name in enumerate(mega_report_df.columns, start=1):
        if 'Talk Time' in col_name:
            time_columns.append(col_idx)
        elif 'Rate' in col_name:
            rate_columns.append(col_idx)
    
    # Apply bold formatting and time format to the time columns
    bold_font = Font(bold=True)
    for col_idx in time_columns:
        # Bold the header
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.font = bold_font
        
        # Bold all data cells in the column and apply time format
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = bold_font
            # Format seconds as [h]:mm:ss
            seconds = cell.value if cell.value else 0
            hours = int(seconds // 3600)
            minutes = int((seconds % 3600) // 60)
            seconds_remainder = int(seconds % 60)
            cell.value = f"{hours:02d}:{minutes:02d}:{seconds_remainder:02d}"
    
    # Apply bold formatting and percentage format to the rate columns
    for col_idx in rate_columns:
        # Bold the header
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.font = bold_font
        
        # Bold all data cells in the column and apply percentage format
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = bold_font
            cell.number_format = '0.00%'  # Format as percentage with 2 decimal places
    
    # Save the formatted workbook
    workbook.save(excel_path)
    
    # Create weekly dataframe for return value (but don't save to CSV)
    weekly_results = []
    current_monday = start_monday
    while current_monday <= current_date:
        week_end = current_monday + timedelta(days=6)  # Sunday
        week_end = min(week_end, current_date)
        
        for agent in unique_agents:
            agent_data = df[df['Effective Agent'] == agent]
            weekly_data = agent_data[(agent_data['Date'] >= current_monday) & 
                                   (agent_data['Date'] <= week_end)]
            
            if len(weekly_data) > 0:
                weekly_metrics = calculate_call_data(weekly_data)
                
                weekly_results.append({
                    'Agent': agent,
                    'Week Start': current_monday.strftime('%Y-%m-%d'),
                    'Week End': week_end.strftime('%Y-%m-%d'),
                    'Total Calls': weekly_metrics['Total Calls'],
                    'Answered Calls': weekly_metrics['Answered Calls'],
                    'Answer Rate': weekly_metrics['Answer Rate'],
                    'Total Talk Time': weekly_metrics['Total Talk Time (seconds)'],
                    'Average Talk Time': weekly_metrics['Average Talk Time (seconds)']
                })
        
        current_monday += timedelta(days=7)
    
    weekly_df = pd.DataFrame(weekly_results)
    
    return {
        'ytd_mtd': ytd_mtd_df,
        'weekly': weekly_df,
        'mega_report': mega_report_df
    }

if __name__ == "__main__":
    # Run the report with default parameters
    interval_calls() 