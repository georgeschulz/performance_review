import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

def channel_stats():
    # Read the CSV file
    df = pd.read_csv('data/Leads-Reporting Export.csv')
    
    # Fill NaN values in Lead Type with "Unknown"
    df['Lead Type'] = df['Lead Type'].fillna('Unknown')
    
    # Convert Close Date to datetime and add Month column
    df['Close Date'] = pd.to_datetime(df['Close Date'], errors='coerce')
    df['Month'] = df['Close Date'].dt.strftime('%B %y')
    
    # Initialize results list
    results = []
    
    # Get unique months and lead types (excluding NaN values)
    months = sorted(df['Month'].dropna().unique(), 
                   key=lambda x: pd.to_datetime(x, format='%B %y'))
    lead_types = df['Lead Type'].dropna().unique()
    
    # Calculate metrics for each month and lead type
    for month in months:
        month_data = df[df['Month'] == month]
        row_data = {'Month': month}
        
        for lead_type in lead_types:
            type_data = month_data[month_data['Lead Type'] == lead_type]
            
            # Calculate metrics
            total_leads = len(type_data)
            disqualified = type_data['Close Status'].str.contains('Disqualified', na=False).sum()
            won_total = (
                (type_data['Close Status'] == 'Won: Recurring').sum() +
                (type_data['Close Status'] == 'Won: One Time').sum()
            )
            won_recurring = (type_data['Close Status'] == 'Won: Recurring').sum()
            
            # Calculate close rate
            lost = type_data['Close Status'].str.contains('Lost', na=False).sum()
            denominator = lost + won_total
            close_rate = (won_total / denominator * 100) if denominator > 0 else 0
            
            # Add to row data with lead type prefix
            prefix = lead_type.replace(' ', '_')
            row_data.update({
                f'{prefix}_Total': total_leads,
                f'{prefix}_Disqualified': disqualified,
                f'{prefix}_Won': won_total,
                f'{prefix}_Won_Recurring': won_recurring,
                f'{prefix}_Close_Rate': f'{close_rate:.2f}%'
            })
        
        results.append(row_data)
    
    # Convert results to DataFrame and save
    results_df = pd.DataFrame(results)
    
    # Create Excel writer
    with pd.ExcelWriter('outputs/Channel Stats.xlsx', engine='openpyxl') as writer:
        # First create a multi-level header DataFrame
        new_columns = pd.MultiIndex.from_tuples([
            ('Month', ''),  # Special handling for Month column
        ])
        
        # Add column groups for each lead type
        for lead_type in df['Lead Type'].unique():
            new_columns = new_columns.append(pd.MultiIndex.from_tuples([
                (lead_type, 'Total Leads'),
                (lead_type, 'Disqualified'),
                (lead_type, 'Won'),
                (lead_type, 'Won Recurring'),
                (lead_type, 'Close Rate')
            ]))
        
        # Create new DataFrame with multi-level columns
        formatted_df = pd.DataFrame(columns=new_columns)
        formatted_df['Month', ''] = results_df['Month']
        
        # Fill in data for each lead type
        for lead_type in df['Lead Type'].unique():
            prefix = lead_type.replace(' ', '_')
            formatted_df[lead_type, 'Total Leads'] = results_df[f'{prefix}_Total']
            formatted_df[lead_type, 'Disqualified'] = results_df[f'{prefix}_Disqualified']
            formatted_df[lead_type, 'Won'] = results_df[f'{prefix}_Won']
            formatted_df[lead_type, 'Won Recurring'] = results_df[f'{prefix}_Won_Recurring']
            formatted_df[lead_type, 'Close Rate'] = results_df[f'{prefix}_Close_Rate']
        
        # Write to Excel with index to avoid MultiIndex columns issue
        formatted_df.to_excel(writer, sheet_name='Channel Stats', index=True)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Channel Stats']
        
        # Hide the index column
        worksheet.column_dimensions['A'].hidden = True
        
        # Define colors for each lead type
        colors = {
            'Thumbtack': 'E2EFDA',  # Light green
            'Inbound': 'B7E1CD',    # Mint green
            'Unknown': 'D9D9D9',    # Light gray
            'Email Lead': 'C9DAF8',  # Light blue
            'Form Fill': 'D9EAD3',   # Pale green
            'Outbound': 'FFD966',    # Gold
            'Sentricon Lead': 'FFF2CC', # Light yellow
            'Other Tech Lead': 'F4CCCC', # Light red
            'Referral': 'D0E0E3',    # Light cyan
            'WTR Lead': 'FBE5D6',     # Light orange,
            'WTR Free Trial Request': 'E6B8AF'  # Light pink/mauve
        }
        
        # Format header rows
        current_col = 2  # Start from column B (since A is hidden index)
        
        # Skip Month column
        current_col += 1
        
        # Format each lead type group
        for lead_type in df['Lead Type'].unique():
            # Each lead type has 5 columns
            start_col = current_col
            end_col = current_col + 4
            
            # Apply color to header cells
            fill = PatternFill(start_color=colors.get(lead_type, 'D9D9D9'),
                             end_color=colors.get(lead_type, 'D9D9D9'),
                             fill_type='solid')
            
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = fill
                cell.font = Font(bold=True, color='000000')
            
            current_col += 5
        
        # Adjust column widths (skip hidden index column)
        for col in range(2, worksheet.max_column + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 15
    
    return formatted_df

if __name__ == "__main__":
    channel_stats()