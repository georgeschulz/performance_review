import pandas as pd
import datetime
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font

def get_fiscal_year_start(current_date=None):
    """
    Calculate the start date of the current fiscal year (September 1st)
    """
    if current_date is None:
        current_date = datetime.datetime.now()
    
    # If current date is before September 1st, fiscal year started last year
    if current_date.month < 9:
        return datetime.datetime(current_date.year - 1, 9, 1)
    # Otherwise fiscal year started this year
    else:
        return datetime.datetime(current_date.year, 9, 1)

def get_month_start(current_date=None):
    """
    Calculate the start date of the current month
    """
    if current_date is None:
        current_date = datetime.datetime.now()
    
    return datetime.datetime(current_date.year, current_date.month, 1)

def calculate_close_rate(data):
    """
    Calculate close rate from a dataframe of leads
    """
    lost = data['Close Status'].str.contains('Lost', na=False).sum()
    won_recurring = (data['Close Status'] == 'Won: Recurring').sum()
    won_one_time = (data['Close Status'] == 'Won: One Time').sum()
    
    denominator = lost + won_recurring + won_one_time
    close_rate = (won_recurring / denominator * 100) if denominator > 0 else 0
    
    return {
        'Lost': lost,
        'Won: Recurring': won_recurring,
        'Won: One Time': won_one_time,
        'Total Closed': denominator,
        'Close Rate': close_rate
    }

def interval_close_rate(beginning_of_time=None, salespeople=[], exclude_channels=[], current_date=None):
    """
    Generate interval-based close rate reports:
    1. Year-to-date and month-to-date close rates
    2. Weekly close rates (Monday to Sunday) with MTD and YTD metrics for each week
    
    Parameters:
    - beginning_of_time: datetime or string, minimum date to include leads from
    - salespeople: list of salespeople to include (empty list means all)
    - exclude_channels: list of lead types to exclude
    - current_date: datetime, date to use as "now" (defaults to today)
    """
    # Set current date if not provided
    if current_date is None:
        current_date = datetime.datetime.now()
    elif isinstance(current_date, str):
        current_date = pd.to_datetime(current_date)
    
    # Read the CSV file
    df = pd.read_csv('weekly_review_data/Leads-Reporting Export.csv')
    
    # Convert Close Date to datetime
    df['Close Date'] = pd.to_datetime(df['Close Date'], errors='coerce')
    
    # Filter by beginning_of_time if provided
    if beginning_of_time:
        if isinstance(beginning_of_time, str):
            beginning_of_time = pd.to_datetime(beginning_of_time)
        df = df[df['Close Date'] >= beginning_of_time]
    
    # Filter by salespeople if provided
    if not salespeople:
        salespeople = df['Salesperson'].unique()
    else:
        df = df[df['Salesperson'].isin(salespeople)]
    
    # Exclude specified channels
    if exclude_channels:
        df = df[~df['Lead Type'].isin(exclude_channels)]
    
    # Calculate fiscal year and month start dates
    fiscal_year_start = get_fiscal_year_start(current_date)
    
    # Initialize results for YTD and MTD report
    ytd_mtd_results = []
    
    # Calculate YTD and MTD metrics for each salesperson
    for sp in salespeople:
        sp_data = df[df['Salesperson'] == sp]
        
        # Year-to-date data
        ytd_data = sp_data[(sp_data['Close Date'] >= fiscal_year_start) & 
                           (sp_data['Close Date'] <= current_date)]
        ytd_metrics = calculate_close_rate(ytd_data)
        
        # Month-to-date data
        month_start = get_month_start(current_date)
        mtd_data = sp_data[(sp_data['Close Date'] >= month_start) & 
                           (sp_data['Close Date'] <= current_date)]
        mtd_metrics = calculate_close_rate(mtd_data)
        
        # Add to results
        ytd_mtd_results.append({
            'Salesperson': sp,
            'YTD Lost': ytd_metrics['Lost'],
            'YTD Won: Recurring': ytd_metrics['Won: Recurring'],
            'YTD Won: One Time': ytd_metrics['Won: One Time'],
            'YTD Total Closed': ytd_metrics['Total Closed'],
            'YTD Close Rate': ytd_metrics['Close Rate'],  # Store as numeric value
            'MTD Lost': mtd_metrics['Lost'],
            'MTD Won: Recurring': mtd_metrics['Won: Recurring'],
            'MTD Won: One Time': mtd_metrics['Won: One Time'],
            'MTD Total Closed': mtd_metrics['Total Closed'],
            'MTD Close Rate': mtd_metrics['Close Rate']  # Store as numeric value
        })
    
    ytd_mtd_df = pd.DataFrame(ytd_mtd_results)
    
    # Initialize results for mega report (weekly + MTD + YTD)
    mega_report_results = []
    
    # Find the Monday of the week containing the fiscal year start date
    # This ensures we start on a Monday
    days_since_monday = fiscal_year_start.weekday()
    start_monday = fiscal_year_start - timedelta(days=days_since_monday)
    
    # Generate weekly periods from start_monday to current_date
    current_monday = start_monday
    while current_monday <= current_date:
        week_end = current_monday + timedelta(days=6)  # Sunday
        
        # If week_end is beyond current_date, cap it at current_date
        week_end = min(week_end, current_date)
        
        # Get the month start for this specific week
        week_month_start = get_month_start(week_end)
        
        # Calculate weekly metrics for each salesperson
        for sp in salespeople:
            sp_data = df[df['Salesperson'] == sp]
            
            # Weekly data (just for this week)
            weekly_data = sp_data[(sp_data['Close Date'] >= current_monday) & 
                                 (sp_data['Close Date'] <= week_end)]
            weekly_metrics = calculate_close_rate(weekly_data)
            
            # YTD data up to and including the end of this week
            ytd_data = sp_data[(sp_data['Close Date'] >= fiscal_year_start) & 
                              (sp_data['Close Date'] <= week_end)]
            ytd_metrics = calculate_close_rate(ytd_data)
            
            # MTD data up to and including the end of this week
            mtd_data = sp_data[(sp_data['Close Date'] >= week_month_start) & 
                              (sp_data['Close Date'] <= week_end)]
            mtd_metrics = calculate_close_rate(mtd_data)
            
            # Add to results if there's any data for the week
            if weekly_metrics['Total Closed'] > 0:
                mega_report_results.append({
                    'Salesperson': sp,
                    'Week Start': current_monday.strftime('%Y-%m-%d'),
                    'Week End': week_end.strftime('%Y-%m-%d'),
                    # Weekly metrics
                    'Weekly Lost': weekly_metrics['Lost'],
                    'Weekly Won: Recurring': weekly_metrics['Won: Recurring'],
                    'Weekly Won: One Time': weekly_metrics['Won: One Time'],
                    'Weekly Total Closed': weekly_metrics['Total Closed'],
                    'Weekly Close Rate': weekly_metrics['Close Rate'],  # Store as numeric value
                    # MTD metrics as of this week (including this week)
                    'MTD Lost': mtd_metrics['Lost'],
                    'MTD Won: Recurring': mtd_metrics['Won: Recurring'],
                    'MTD Won: One Time': mtd_metrics['Won: One Time'],
                    'MTD Total Closed': mtd_metrics['Total Closed'],
                    'MTD Close Rate': mtd_metrics['Close Rate'],  # Store as numeric value
                    # YTD metrics as of this week (including this week)
                    'YTD Lost': ytd_metrics['Lost'],
                    'YTD Won: Recurring': ytd_metrics['Won: Recurring'],
                    'YTD Won: One Time': ytd_metrics['Won: One Time'],
                    'YTD Total Closed': ytd_metrics['Total Closed'],
                    'YTD Close Rate': ytd_metrics['Close Rate']  # Store as numeric value
                })
        
        # Move to next Monday
        current_monday += timedelta(days=7)
    
    # Convert mega report results to DataFrame and save as CSV
    mega_report_df = pd.DataFrame(mega_report_results)
    
    # Format percentages for CSV output
    csv_df = mega_report_df.copy()
    for col in csv_df.columns:
        if 'Close Rate' in col:
            csv_df[col] = csv_df[col].apply(lambda x: f"{x:.2f}%")
        
    # Also save as Excel with minimal formatting (just bold for close rate columns)
    excel_path = 'weekly_outputs/Close Rate Report.xlsx'
    mega_report_df.to_excel(excel_path, index=False, engine='openpyxl')
    
    # Apply formatting to the Excel file
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook.active
    
    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        # Find the maximum length in the column
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Set the column width (with some padding)
        adjusted_width = max_length + 4
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Find the column indices for close rate columns
    close_rate_columns = []
    for col_idx, col_name in enumerate(mega_report_df.columns, start=1):
        if 'Close Rate' in col_name:
            close_rate_columns.append(col_idx)
    
    # Apply bold formatting and percentage format to the close rate columns
    bold_font = Font(bold=True)
    for col_idx in close_rate_columns:
        # Bold the header
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.font = bold_font
        
        # Bold all data cells in the column and apply percentage format
        for row_idx in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = bold_font
            cell.number_format = '0.00"%"'  # Format as percentage with 2 decimal places
    
    # Save the formatted workbook
    workbook.save(excel_path)
    
    # Also keep the original weekly report for backward compatibility
    weekly_results = []
    current_monday = start_monday
    while current_monday <= current_date:
        week_end = current_monday + timedelta(days=6)  # Sunday
        week_end = min(week_end, current_date)
        
        for sp in salespeople:
            sp_data = df[df['Salesperson'] == sp]
            weekly_data = sp_data[(sp_data['Close Date'] >= current_monday) & 
                                 (sp_data['Close Date'] <= week_end)]
            weekly_metrics = calculate_close_rate(weekly_data)
            
            if weekly_metrics['Total Closed'] > 0:
                weekly_results.append({
                    'Salesperson': sp,
                    'Week Start': current_monday.strftime('%Y-%m-%d'),
                    'Week End': week_end.strftime('%Y-%m-%d'),
                    'Lost': weekly_metrics['Lost'],
                    'Won: Recurring': weekly_metrics['Won: Recurring'],
                    'Won: One Time': weekly_metrics['Won: One Time'],
                    'Total Closed': weekly_metrics['Total Closed'],
                    'Close Rate': weekly_metrics['Close Rate']  # Store as numeric value
                })
        
        current_monday += timedelta(days=7)
    
    weekly_df = pd.DataFrame(weekly_results)
    
    # Format percentages for the weekly report CSV
    weekly_df_csv = weekly_df.copy()
    weekly_df_csv['Close Rate'] = weekly_df_csv['Close Rate'].apply(lambda x: f"{x:.2f}%")
    
    # Format percentages for the YTD/MTD report CSV
    ytd_mtd_df_csv = ytd_mtd_df.copy()
    ytd_mtd_df_csv['YTD Close Rate'] = ytd_mtd_df_csv['YTD Close Rate'].apply(lambda x: f"{x:.2f}%")
    ytd_mtd_df_csv['MTD Close Rate'] = ytd_mtd_df_csv['MTD Close Rate'].apply(lambda x: f"{x:.2f}%")
    
    return {
        'ytd_mtd': ytd_mtd_df,
        'weekly': weekly_df,
        'mega_report': mega_report_df
    }

